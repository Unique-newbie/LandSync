"""Report Generation Service"""

import os
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional
import json


class ReportGenerator:
    """
    Generate reports in various formats.
    
    Formats:
    - PDF: Using reportlab
    - Excel: Using openpyxl
    - CSV: Standard CSV
    """
    
    REPORTS_DIR = "uploads/reports"
    
    def __init__(self):
        """Initialize report generator"""
        os.makedirs(self.REPORTS_DIR, exist_ok=True)
    
    def generate(
        self,
        data: Dict[str, Any],
        report_type: str,
        format: str,
        report_id: uuid.UUID,
        title: str
    ) -> str:
        """
        Generate a report file.
        
        Args:
            data: Report data
            report_type: Type of report
            format: Output format (pdf, excel, csv)
            report_id: Unique report ID
            title: Report title
            
        Returns:
            Path to generated file
        """
        
        if format == "pdf":
            return self._generate_pdf(data, report_type, report_id, title)
        elif format == "excel":
            return self._generate_excel(data, report_type, report_id, title)
        elif format == "csv":
            return self._generate_csv(data, report_type, report_id, title)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def _generate_pdf(
        self,
        data: Dict[str, Any],
        report_type: str,
        report_id: uuid.UUID,
        title: str
    ) -> str:
        """Generate PDF report"""
        
        file_path = os.path.join(self.REPORTS_DIR, f"{report_id}.pdf")
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
            from reportlab.lib.units import inch
            
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30
            )
            elements.append(Paragraph(title, title_style))
            
            # Date
            elements.append(Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 20))
            
            # Content based on report type
            if report_type == "reconciliation":
                self._add_reconciliation_content(elements, data, styles)
            elif report_type == "mismatch":
                self._add_mismatch_content(elements, data, styles)
            elif report_type == "summary":
                self._add_summary_content(elements, data, styles)
            
            doc.build(elements)
            
        except ImportError:
            # Fallback: Generate simple text file with .pdf extension
            with open(file_path, 'w') as f:
                f.write(f"# {title}\n\n")
                f.write(f"Generated: {datetime.now().isoformat()}\n\n")
                f.write(json.dumps(data, indent=2, default=str))
        
        return file_path
    
    def _generate_excel(
        self,
        data: Dict[str, Any],
        report_type: str,
        report_id: uuid.UUID,
        title: str
    ) -> str:
        """Generate Excel report"""
        
        file_path = os.path.join(self.REPORTS_DIR, f"{report_id}.xlsx")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Title
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            row = 4
            
            if report_type == "reconciliation":
                # Headers
                headers = ["Parcel ID", "Record ID", "Score", "Status", "Confidence"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="CCCCCC")
                
                row += 1
                for match in data.get("matches", []):
                    ws.cell(row=row, column=1, value=match.get("parcel_id"))
                    ws.cell(row=row, column=2, value=match.get("record_id"))
                    ws.cell(row=row, column=3, value=match.get("score"))
                    ws.cell(row=row, column=4, value=match.get("status"))
                    ws.cell(row=row, column=5, value=match.get("confidence"))
                    row += 1
                    
            elif report_type == "mismatch":
                headers = ["Parcel Plot ID", "Parcel Owner", "Parcel Area",
                          "Record Plot ID", "Record Owner", "Record Area", "Score"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                
                row += 1
                for m in data.get("mismatches", []):
                    ws.cell(row=row, column=1, value=m.get("parcel_plot_id"))
                    ws.cell(row=row, column=2, value=m.get("parcel_owner"))
                    ws.cell(row=row, column=3, value=m.get("parcel_area"))
                    ws.cell(row=row, column=4, value=m.get("record_plot_id"))
                    ws.cell(row=row, column=5, value=m.get("record_owner"))
                    ws.cell(row=row, column=6, value=m.get("record_area"))
                    ws.cell(row=row, column=7, value=m.get("score"))
                    row += 1
                    
            elif report_type == "summary":
                ws.cell(row=row, column=1, value="Total Parcels")
                ws.cell(row=row, column=2, value=data.get("parcels", 0))
                row += 1
                ws.cell(row=row, column=1, value="Total Records")
                ws.cell(row=row, column=2, value=data.get("records", 0))
                row += 1
                ws.cell(row=row, column=1, value="Total Matches")
                ws.cell(row=row, column=2, value=data.get("matches", 0))
                row += 1
                ws.cell(row=row, column=1, value="Average Score")
                ws.cell(row=row, column=2, value=data.get("average_score", 0))
            
            wb.save(file_path)
            
        except ImportError:
            # Fallback to CSV
            return self._generate_csv(data, report_type, report_id, title)
        
        return file_path
    
    def _generate_csv(
        self,
        data: Dict[str, Any],
        report_type: str,
        report_id: uuid.UUID,
        title: str
    ) -> str:
        """Generate CSV report"""
        
        import csv
        
        file_path = os.path.join(self.REPORTS_DIR, f"{report_id}.csv")
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            if report_type == "reconciliation":
                writer.writerow(["Parcel ID", "Record ID", "Score", "Status", "Confidence"])
                for match in data.get("matches", []):
                    writer.writerow([
                        match.get("parcel_id"),
                        match.get("record_id"),
                        match.get("score"),
                        match.get("status"),
                        match.get("confidence")
                    ])
                    
            elif report_type == "mismatch":
                writer.writerow([
                    "Parcel Plot ID", "Parcel Owner", "Parcel Area",
                    "Record Plot ID", "Record Owner", "Record Area", "Score"
                ])
                for m in data.get("mismatches", []):
                    writer.writerow([
                        m.get("parcel_plot_id"),
                        m.get("parcel_owner"),
                        m.get("parcel_area"),
                        m.get("record_plot_id"),
                        m.get("record_owner"),
                        m.get("record_area"),
                        m.get("score")
                    ])
                    
            elif report_type == "summary":
                writer.writerow(["Metric", "Value"])
                writer.writerow(["Total Parcels", data.get("parcels", 0)])
                writer.writerow(["Total Records", data.get("records", 0)])
                writer.writerow(["Total Matches", data.get("matches", 0)])
                writer.writerow(["Average Score", data.get("average_score", 0)])
        
        return file_path
    
    def _add_reconciliation_content(
        self,
        elements: list,
        data: Dict[str, Any],
        styles
    ):
        """Add reconciliation report content"""
        
        from reportlab.platypus import Paragraph, Table, TableStyle
        from reportlab.lib import colors
        
        elements.append(Paragraph(f"Total Matches: {data.get('total', 0)}", styles['Heading2']))
        
        if data.get("matches"):
            table_data = [["Parcel ID", "Record ID", "Score", "Status"]]
            for match in data["matches"][:100]:  # Limit to 100 rows
                table_data.append([
                    match.get("parcel_id", "")[:20],
                    match.get("record_id", "")[:20],
                    str(match.get("score", 0)),
                    match.get("status", "")
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
    
    def _add_mismatch_content(
        self,
        elements: list,
        data: Dict[str, Any],
        styles
    ):
        """Add mismatch report content"""
        
        from reportlab.platypus import Paragraph
        
        elements.append(Paragraph(
            f"Total Mismatches: {data.get('total', 0)}",
            styles['Heading2']
        ))
        
        for m in data.get("mismatches", [])[:20]:
            elements.append(Paragraph(
                f"• {m.get('parcel_plot_id')} vs {m.get('record_plot_id')} "
                f"(Score: {m.get('score')})",
                styles['Normal']
            ))
    
    def _add_summary_content(
        self,
        elements: list,
        data: Dict[str, Any],
        styles
    ):
        """Add summary report content"""
        
        from reportlab.platypus import Paragraph
        
        elements.append(Paragraph(f"Parcels: {data.get('parcels', 0)}", styles['Normal']))
        elements.append(Paragraph(f"Records: {data.get('records', 0)}", styles['Normal']))
        elements.append(Paragraph(f"Matches: {data.get('matches', 0)}", styles['Normal']))
        elements.append(Paragraph(f"Average Score: {data.get('average_score', 0)}", styles['Normal']))
